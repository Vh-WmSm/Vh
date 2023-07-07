import os
import sys

import numpy
import datetime
import openpyxl
import public_tools
from openpyxl.chart import *
from openpyxl.styles import Alignment


class LiuHongLiang_DataProcessing:
    def __init__(self):
        self.input_path = None
        self.output_path = None
        self.excel_path_list = []  # 递归获取input_path中的所有Excel文件路径
        self.Excel_sort_key = 0  # 人为检查Excel文件路径排序是否正确，若不正确则先调整文件夹的名字
        self.start_date = None  # 开始日期（用于命名表格名称）
        self.end_date = None  # 结束日期
        self.first_Excel_key = 0  # 是否遍历第一个Excel
        self.integration_wb = None  # Excel工作簿（整合）
        self.integration_ws = None
        self.prjTuple = None  # 获取每一列对象，封装成((列1), (列2), (列3)...)，其中，如对(列1)[0].value可获取第一列的第一个单元格内容
        self.CH_list = []  # 如CH1
        self.CH_name_list = []  # 如S14448
        self.CH_count = 0  # CH通道总个数
        self.group_count = 1  # 组编号（一个小时）
        self.group_data_list = []
        self.group_data_list_after_mean = []

    def Gui(self):
        desktop_path = public_tools.My_Tool.get_desktop_path()
        title_name_lis = ['请选择要处理的文件夹:', 'Excel文件输出位置(默认桌面):']
        windows_name = '光源测试数据处理整合小工具'
        button_name = '开始处理'
        default_value_lis = ['folder', desktop_path]
        self.GUI = public_tools.PySimpleGUI_Tool(title_name_lis, windows_name, button_name, default_value_lis,
                                                 in_size=25,
                                                 title_size=22)

    def Gui_data_return(self):
        self.input_path, self.output_path = self.GUI.gui_windows()

    def __recursive_function(self, path, last_path):
        os.chdir(path)
        lis = os.listdir()
        for li in lis:
            # 是目录
            if os.path.isdir(li):
                self.__recursive_function(path + f'\\{li}', path)
            # 是文件
            else:
                if '.' in li:
                    # 分割
                    sp = li.rsplit('.', 1)
                    # 获取文件名
                    name = sp[0]
                    # 获取后缀
                    suffix = sp[1]
                    if suffix in ('xlsx', 'xls'):
                        # 若是xls文件，则改名，且更新li
                        if 'xls' == suffix:
                            os.rename(li, name + '.xlsx')
                            li = name + '.xlsx'
                        # 记录该Excel文件路径
                        self.excel_path_list.append(path + f'\\{li}')
        # 返回上一级
        os.chdir(last_path)

    def get_all_excel_path_recursively(self):
        text = '以下是递归获取到目标文件夹下的所有Excel\n注意：\n1.程序会按照下方的排列顺序进行数据获取\n2.若有不需要整合的Excel表，把不需要的表移出文件夹后重新获取列表\n3.时间和“功率测试结果”之间必须有一个空格且只有这一个空格，时间格式必须年月日时分秒且“-”分隔，若排序错误或程序出错，请检查此问题\n4.要保证第一个Excel的光功率计通道数和SLD光源编号顺序与后面的Excel保持一致\n5.无论是否符合要求，请先按OK\n'
        self.__recursive_function(self.input_path, self.input_path)
        # Excel列表排序
        self.excel_path_list.sort(key=lambda x: datetime.datetime.strptime(x.rsplit('\\', 1)[1].split(' ', 1)[0],
                                                                           "%Y-%m-%d-%H-%M-%S"))  # str转为时间对象进行比较
        for ex_li in self.excel_path_list:
            text += ex_li.rsplit('\\', 1)[1] + '\n'
        self.GUI.popup_scrolled(text, title='提示')
        if self.GUI.popup_yes_no('符合要求吗？') == 'Yes':
            # 修改标记值
            self.Excel_sort_key = 1

    def integration_pyxl(self):
        self.integration_wb = openpyxl.Workbook()
        self.integration_ws = self.integration_wb.active
        self.integration_ws.title = f'{self.start_date}-{self.end_date}光源测试结果'
        # 设置列宽
        self.integration_ws.column_dimensions['a'].width = 17  # 标题/组号列
        for CH_column in range(1, self.CH_count + 1):
            self.integration_ws.column_dimensions[f'{chr(CH_column + 65)}'].width = 15
        # 更改当前聚焦的行
        self.integration_ws._current_row = 9
        row10 = [None]
        row10.extend([f'CH{i}' for i in range(1, self.CH_count + 1)])
        self.integration_ws.append(row10)
        for after_mean_data_li in self.group_data_list_after_mean:
            # 在第11行开始插入数据
            self.integration_ws.append(after_mean_data_li)
        # 保存半成品，等待手动处理无效数据
        self.integration_wb.save(f'{self.output_path}\\{self.start_date}-{self.end_date}光源测试结果.xlsx')
        self.integration_wb.close()

    def manual_operation(self, for_times):
        GUI = public_tools.PySimpleGUI_Tool()
        GUI.popup_scrolled(
            '请手动处理好无效数据，完成后保存关闭Excel再按OK！\n处理后会产生待定结果(1)、(2)...，若不满意可以继续修改',
            title='提示')
        self.integration_wb = openpyxl.load_workbook(
            f'{self.output_path}\\{self.start_date}-{self.end_date}光源测试结果.xlsx')
        self.integration_ws = self.integration_wb.active
        max_row = self.integration_ws.max_row
        max_column = self.integration_ws.max_column
        # 画折线图，定义LineChart对象
        chart = LineChart()
        # 定义x坐标的数据
        x_data = Reference(self.integration_ws, min_col=1, min_row=11, max_col=1, max_row=max_row)
        # 定义作图数据列
        data = Reference(self.integration_ws, min_col=2, min_row=10, max_col=max_column, max_row=max_row)
        # 添加数据到chart（from_rows=False说明是以列数据来画图，titles_from_data=False说明不画图例，若为True则以第一行文字写图例，这时上方的data处应min_row=1）
        chart.add_data(data, from_rows=False, titles_from_data=True)
        # 添加x坐标数据
        chart.set_categories(x_data)
        # 定义纵坐标名称
        chart.y_axis.title = '平均值'
        # 定义横坐标名称
        chart.x_axis.title = '组'
        # 定义折线图标题
        chart.title = f'{self.start_date}-{self.end_date}光源测试结果整合折线图'
        # 把折线图画在E12
        chart.anchor = 'e12'
        # 设置折线图大小
        chart.height = 15
        chart.width = 25
        self.integration_ws.add_chart(chart)
        # 更改当前聚焦的行（设置为1，让下一个append在第二行插入（因为模板的第一行是空的））
        self.integration_ws._current_row = 1
        # 获取当前最大行
        max_row = self.integration_ws.max_row
        # 定义title_and_init_list
        title_and_init_list = []
        row2 = ['光功率计通道']
        row2.extend([i for i in range(1, self.CH_count + 1)])
        title_and_init_list.append(row2)
        row3 = ['SLD光源编号']
        row3.extend(self.CH_name_list)
        title_and_init_list.append(row3)
        row4 = ['线性系数uW/h']
        row4.extend([f'=LINEST({chr(x + 65)}11:{chr(x + 65)}{max_row},$A$11:$A${max_row},1)' for x in
                     range(1, self.CH_count + 1)])
        title_and_init_list.append(row4)
        row5 = ['不稳定度/%']
        row5.extend(
            [f'=2*STDEV.S({chr(x + 65)}11:{chr(x + 65)}{max_row})/AVERAGE({chr(x + 65)}11:{chr(x + 65)}{max_row})*100'
             for x in range(1, self.CH_count + 1)])
        title_and_init_list.append(row5)
        row6 = ['变化率/%']
        row6.extend([
            f'=0.5*(MAX({chr(x + 65)}11:{chr(x + 65)}{max_row})-MIN({chr(x + 65)}11:{chr(x + 65)}{max_row}))/AVERAGE({chr(x + 65)}11:{chr(x + 65)}{max_row})*100'
            for x in range(1, self.CH_count + 1)])
        title_and_init_list.append(row6)
        row7 = ['相对变化率/(h·%)']
        row7.extend([f'={chr(x + 65)}4/{chr(x + 65)}11' for x in range(1, self.CH_count + 1)])
        title_and_init_list.append(row7)
        row8 = ['光源等级（TEC关）']
        row8.extend([f'=IF(ABS({chr(x + 65)}7)>0.0001,"C",IF(ABS({chr(x + 65)}7)<0.00005,"A","B"))' for x in
                     range(1, self.CH_count + 1)])
        title_and_init_list.append(row8)
        for title_and_init_li in title_and_init_list:
            self.integration_ws.append(title_and_init_li)
        # 获取最大列数
        max_column = self.integration_ws.max_column
        # 所有数据居中
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                self.integration_ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
        self.integration_wb.save(f'{self.output_path}\\(待定结果{for_times}){self.start_date}-{self.end_date}光源测试结果.xlsx')
        self.integration_wb.close()
        # 死循环 —— 等待popup_yes_no返回判断结果
        while True:
            if self.GUI.popup_yes_no('是否确定此版本？是则按Yes，不满意则按No继续修改', title='提示') == 'Yes':
                sys.exit()
            else:
                break

    def each_pyxl(self):
        for excel_path in self.excel_path_list:
            # 遍历获取每一个Excel表
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            # 获取当前Excel表的最大行数
            ws_max_row = ws.max_row
            # 获取开始日期和结束日期（因为已排序，所以第一个Excel的A4单元格肯定是最小日期，最后一个Excel的A列最后一行的单元格肯定是最大日期）
            if excel_path == self.excel_path_list[0]:
                start_sp = ws['a4'].value.split('-')
                self.start_date = f'{int(start_sp[1])}月{int(start_sp[2])}日'
            elif excel_path == self.excel_path_list[-1]:
                end_sp = ws[f'a{ws_max_row}'].value.split('-')
                self.end_date = f'{int(end_sp[1])}月{int(end_sp[2])}日'
            # 若是第一个Excel，先获取所有CH通道和名称
            if self.first_Excel_key == 0:
                # 获取所有列
                self.prjTuple = tuple(ws.columns)
                # 从第三列开始获取
                for idx in range(2, len(self.prjTuple), 3):
                    CH = self.prjTuple[idx][0].value
                    CH_name = self.prjTuple[idx][1].value
                    if CH is not None:
                        self.CH_list.append(CH)
                    if CH_name is not None:
                        self.CH_name_list.append(CH_name)
                # 获取CH的个数
                self.CH_count = len(self.CH_list)
            for row in range(4, ws_max_row + 2):
                for CH_column in range(3, self.CH_count * 3 + 1, 3):
                    # 每一行的第一次遍历，先获取对应的时间
                    if CH_column == 3:
                        time_cell_data = ws.cell(row=row, column=1).value
                        if time_cell_data is not None:
                            # 获取时间中的“时”
                            new_time = time_cell_data.rsplit('-', 1)[1].split(':')[0]
                    # 若是第一个Excel表，则把last_time初始化，用于比较“时”是否改变
                    if self.first_Excel_key == 0:
                        last_time = new_time
                        # 切换标记
                        self.first_Excel_key = 1
                    if row != ws_max_row + 1:
                        # 若新获取的“时”和上一个“时”相同，则把值加入该组列表
                        if last_time == new_time:
                            # 最终形成列表[[CH1的值1, CH1的值2...], [CH2的值1, CH2的值2...], ...]
                            # 组建内部列表（当列表内的列表数不足self.CH_count时，就运行此if）
                            if len(self.group_data_list) < self.CH_count:
                                self.group_data_list.append([ws.cell(row=row, column=CH_column).value])
                            else:
                                self.group_data_list[CH_column // 3 - 1].append(ws.cell(row=row, column=CH_column).value)
                        else:
                            for group_data_li in enumerate(self.group_data_list):
                                if group_data_li[0] == 0:
                                    # 先写入该组编号（编号从1开始）
                                    lis = [self.group_count]
                                # 求平均乘以1000加入列表
                                lis.append(numpy.mean(group_data_li[1]) * 1000)
                            self.group_data_list_after_mean.append(lis)
                            # 自增编号变量
                            self.group_count += 1
                            # 清空上一组的列表
                            self.group_data_list.clear()
                            # 若非该表的最后一行的第一列单元格None
                            # 由于进入这个else的原因是当前的“时”与上一个“时”不相同，说明此时遍历到了下一组的第一行的CH1单元格，所以self.group_data_list列表需清空，然后把当前单元格写入列表
                            self.group_data_list.append([ws.cell(row=row, column=CH_column).value])
                            # 更新last_time时间为当前组的时间
                            last_time = new_time
                    # 若遍历到当前Excel的最大行的下一个空白行（到达边界，此时有两种情况：1.打开下一个能头尾接上的Excel。2.后面没有了Excel表，这是最后一个）
                    else:
                        # 不是最后一个Excel
                        if excel_path != self.excel_path_list[-1]:
                            pass
                        # 最后一个Excel，把截至此的self.group_data_list作为一个组处理后写入列表
                        else:
                            for group_data_li in enumerate(self.group_data_list):
                                if group_data_li[0] == 0:
                                    # 先写入该组编号（编号从1开始）
                                    lis = [self.group_count]
                                # 求平均乘以1000加入列表
                                lis.append(numpy.mean(group_data_li[1]) * 1000)
                            self.group_data_list_after_mean.append(lis)
                        # 由于此为最后一个空行，所以不需获取各列CH通道数据了，直接break跳出列的循环，然后跳出列的循环后，由于已经到达max_row + 1，所以行的循环也会自动跳出，然后打开下一个Excel
                        break
            wb.close()


if __name__ == '__main__':
    lhl_dp = LiuHongLiang_DataProcessing()
    lhl_dp.Gui()
    while lhl_dp.Excel_sort_key == 0:
        lhl_dp.Gui_data_return()
        lhl_dp.get_all_excel_path_recursively()
    lhl_dp.each_pyxl()
    lhl_dp.integration_pyxl()
    # 给多次手动更改的机会
    for for_times in range(1, 100):
        lhl_dp.manual_operation(for_times)
