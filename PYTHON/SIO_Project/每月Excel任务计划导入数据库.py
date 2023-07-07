import openpyxl
import DataBase
import public_tools


class SIO_DataBase_Insert:
    def __init__(self):
        self.db = DataBase.DataBase(host='localhost')  # 创建数据库类对象
        self.tb_task_title = None
        self.tb_staff_datas = []
        self.tb_task_datas = []
        self.project_admin_name_list = []

    def get_info_by_task_excel(self):
        desktop_path = public_tools.My_Tool.get_desktop_path()
        title_name_lis = ['Excel文件地址:', 'Excel文件名(只支持.xlsx):', '从第几行开始:', '第几行结束:',
                          '项目名称/经理所在列:', '结果定义所在列:', '任务负责人所在列:']
        windows_name = '月任务导入数据库小工具'
        button_name = '开始导入'

        default_value_lis = [desktop_path, '1.xlsx', '4', '', 'b', 'd', 'i']
        GUI = public_tools.PySimpleGUI_Tool(title_name_lis, windows_name, button_name, default_value_lis,
                                            file_suffix=['.xlsx'], in_size=25, suffix_in_size=25)
        path, file_name, start_row, end_row, project_start_column, description_start_column, principal_start_column = GUI.gui_windows()
        wb = openpyxl.load_workbook(path + '\\' + file_name)
        ws = wb.active

        # 从Excel的标题获取表的时间作为新任务表的表名
        Excel_title = ws['a1'].value
        self.tb_task_title = 'tb_task_'
        for s in Excel_title:
            if s.isdigit():
                self.tb_task_title += s
        # 数据写入数据库的数据列表（内部包含多个元组）
        self.tb_task_datas = []
        # 从开始行开始遍历Excel
        for i in range(int(start_row), int(end_row) + 1):
            # 获取项目名称
            project_info = ws[f'{project_start_column}{i}'].value
            if (project_info is not None):
                if '项目经理：' in project_info:
                    # 以“：”为分隔，从右往左仅分隔一次，防止有多个“：”
                    sp = project_info.rsplit('：', 1)
                    # 有“项目经理：”字样，且被括号括住
                    if project_info[project_info.find('项目经理') - 1] in ('(', '（'):
                        project_name = sp[0][:-5]
                        project_admin_name = sp[1][:-1]
                    # 有“项目经理：”字样，且没有被括号括住
                    else:
                        project_name = sp[0][:-4]
                        project_admin_name = sp[1]
                # 没有“项目经理：”字样，只写了个名字，但名字被括号括住的情况
                elif '（' in project_info:
                    sp = project_info.rsplit('（', 1)
                    project_name = sp[0]
                    project_admin_name = sp[1][:-1]
                else:
                    project_name = project_info
                    project_admin_name = None
                # 把项目经理名字写入列表
                if project_admin_name is not None:
                    self.project_admin_name_list.append(project_admin_name)

                # 去除项目名字内的换行符
                project_name_ = ''
                for s in project_name:
                    if s not in ('\n', '\xa0'):
                        project_name_ += s
                # 去除项目名字最后的句号
                if (project_name_[-1] == '。'):
                    project_name_ = project_name_[:-1]
                # 去除项目名字后面可能多余的空格
                project_name_ = project_name_.strip()
            # 获取结果定义
            task_description = ws[f'{description_start_column}{i}'].value
            # 若该格结果定义非空，去除结果定义中可能的换行符和前后的空格
            if task_description is not None:
                task_description = ''.join(task_description.split('\n')).strip()
                # 去除结果定义最后可能存在的句号
                if task_description[-1] == '。':
                    task_description = task_description[:-1]
            # 获取该结果定义的负责人
            task_principal_name_temp = ws[f'{principal_start_column}{i}'].value
            if task_principal_name_temp is not None:
                task_principal_name = task_principal_name_temp
            # 把tb_task数据追加到任务列表中
            self.tb_task_datas.append(
                (project_name_, project_admin_name, task_description, task_principal_name))

            # 把负责人名字唯一追加到员工表
            if [task_principal_name] not in self.tb_staff_datas and task_principal_name is not None:
                self.tb_staff_datas.append([task_principal_name])
            # 把项目经理名字唯一追加到员工表
            if [project_admin_name] not in self.tb_staff_datas and project_admin_name is not None:
                self.tb_staff_datas.append([project_admin_name])
        for staff in self.tb_staff_datas:
            # 该员工是项目经理
            if staff[0] in self.project_admin_name_list:
                staff.append(2)  # self.tb_project_admin_datas是[['员工1', 2], ['员工2'], ...]这样的列表
        # 关闭Excel文件
        wb.close()

    def create_and_insert_tb_task_by_month(self):
        # 若存在当月的任务计划表，则先删除（预防重新导入的可能）
        self.db.drop_table_if_exists(self.tb_task_title)
        # 若不存在当月的任务计划表，则创建
        self.db.create_task_table_if_not_exists(self.tb_task_title)
        # 遍历self.tb_task_datas，获取每行元祖数据并插入
        for task_data in self.tb_task_datas:
            self.db.insert(self.tb_task_title, task_data, data_num=4, first_id=True)
        # 提交事务
        self.db.commit()

    def insert_department_admin_staff_first(self):
        # 若员工表不存在，先创建
        self.db.create_staff_if_not_exists()
        department_admin_datas = [['肖浩', 3], ['刘东伟', 3], ['刘博阳', 3], ['李建光', 3], ['廉幸刚', 3]]
        for department_admin_data in department_admin_datas:
            if self.db.select_tb_staff_by_name(department_admin_data[0]) is False:
                self.db.insert_staff(department_admin_data)


    def insert_tb_staff(self):
        # 插入员工信息
        for staff in self.tb_staff_datas:
            # 判断该员工是否已在数据库
            if self.db.select_tb_staff_by_name(staff[0]) is False:
                self.db.insert_staff(staff)
        # 提交事务
        self.db.commit()


if __name__ == '__main__':
    ins = SIO_DataBase_Insert()
    # 获取excel数据
    ins.get_info_by_task_excel()
    # 把相应数据创建月任务表且导入，表根据Excel标题动态命名为例如：tb_task_2305
    ins.create_and_insert_tb_task_by_month()
    # 先插入部门经理信息
    ins.insert_department_admin_staff_first()
    # 唯一插入员工信息
    ins.insert_tb_staff()
    # 关闭数据库
    ins.db.close()
