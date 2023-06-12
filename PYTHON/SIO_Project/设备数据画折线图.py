import os
import openpyxl
import public_tools
from openpyxl.chart import *
from openpyxl.styles import Alignment


class line_chart:
    def init(self):
        self.integration_ws = None
        self.integration_wb = None
        desktop_path = public_tools.My_Tool.get_desktop_path()
        title_name_lis = ['以时间区间命名的文件夹(内放xls/xlsx):', '请选择模式:',
                          '结果文件生成位置:']
        windows_name = '设备数据折线图小工具'
        button_name = '开始'
        default_value_lis = [desktop_path,
                             ['各个设备生成各自折线图', '各个设备折线图整合取绝对误差', '前两个选项都执行'],
                             desktop_path]
        GUI = public_tools.PySimpleGUI_Tool(title_name_lis, windows_name, button_name, default_value_lis)
        self.origin_path, self.mode, self.target_path = GUI.gui_windows()
        self.mode_key = ''
        # 从路径地址获取时间区间（用GUI获取的地址的斜杠是/而不是\了，所以用/分隔）
        self.during_time = self.origin_path.split('/')[-1]
        print(self.origin_path)
        print(self.during_time)
        # 工作目录设定为该文件夹
        os.chdir(self.origin_path)
        # 获取该文件夹的所有文件列表
        lis = os.listdir()
        # lis按文件名（不算后缀）从小到大排序 —— 因为文件名即需延时的时间，从小到大排序可以让延时最小的在前，方便后面的整合时的时间对齐操作
        lis.sort(key=lambda x: x.split('.')[0])
        # 初始化处理后的文件名列表
        self.file_name_lis = []
        # 处理后的文件名
        self.file_name = None
        # 标记：第一个遍历的Excel文件
        self.first_file_key = False
        # 标记：Excel文件是否被遍历完的key
        self.last_file_key = False
        # 数数有多少个xls或xlsx文件
        self.xls_count = 0
        # 遍历lis
        for li in lis:
            if li == lis[0]:
                self.first_file_key = True
            else:
                self.first_file_key = False
            if li == lis[-1]:
                self.last_file_key = True
            suffix = li.split('.')[1]
            self.time_add = int(li.split('.')[0])
            # 找到xls的文件
            if suffix in ('xls', 'xlsx'):
                self.xls_count += 1
                self.name = li
                # 重命名后缀为xlsx（因为openpyxl不能打开xls）
                if suffix == 'xls':
                    os.rename(li, li + 'x')
                    self.name = li + 'x'
                # 打开文件
                self.wb = openpyxl.load_workbook(self.origin_path + '\\' + self.name)
                # 获取sheet1
                self.ws = self.wb.active
                # 获取行数
                max_row = self.ws.max_row
                # 获取最后一行时间（除去秒数）存入tempMinRow，若有与它相同的分钟数，则删除该行
                tempMinRow = self.ws[f'f{max_row}'].value.rsplit(':', 1)[0]
                # for循环遍历当前Excel表，去除相同分钟行，且值一列换为int型（从倒数最后一行开始往上遍历，一直到第二行结束）
                for i in range(max_row, 1, -1):
                    try:
                        # 获取当前分钟行
                        nowMinRow = self.ws[f'f{i}'].value.rsplit(':', 1)[0]
                        # 找到相同的分钟行
                        if tempMinRow == nowMinRow and i != max_row:
                            # 删除该行，保留最早测的分钟行
                            self.ws.delete_rows(i)
                        # 遍历到了下一分钟行
                        else:
                            # 更新tempMinRow
                            tempMinRow = nowMinRow
                            # 把值一列的str类型改为int类型（int类型才可作出折线图）
                            self.ws[f'G{i}'] = int(self.ws[f'G{i}'].value)
                            # 时间处理
                            self.ws[f'f{i}'] = self.time_disposal(self.ws[f'f{i}'].value)
                    except:
                        pass
                # 让设备名-变量名称-从机名称为保存的文件名
                _file_name = self.ws['a2'].value + '-' + self.ws['c2'].value + '-' + self.ws['e2'].value
                self.file_name = _file_name + f' +{self.time_add}min已同步'
                # 把各个文件名写入列表，方便整合时写折线图标题
                self.file_name_lis.append(_file_name)
                if self.mode == '各个设备生成各自折线图':
                    self.mode_key = '1'
                    self.single_mapping()
                elif self.mode == '各个设备折线图整合取绝对误差':
                    self.mode_key = '2'
                    self.table_integration()
                elif self.mode == '前两个选项都执行':
                    self.mode_key = '3'
                    self.single_mapping()
                    self.table_integration()
        if self.mode_key == '1':
            self.wb.close()
        elif self.mode_key in ('2', '3'):
            self.wb.close()
            self.integration_wb.save(self.target_path + '\\IntegrationFile.xlsx')
            self.integration_wb.close()

    def single_mapping(self):
        # 获取初始化时已处理完成后的max_row（由于可能会删除某些相同分钟行，所以max_row需重新获取）
        max_row = self.ws.max_row
        # 设置列宽（A、E、F列文字太多显示不全，需调整列宽）
        self.ws.column_dimensions['a'].width = 20  # 联网设备列
        self.ws.column_dimensions['e'].width = 15  # 从机名称列
        self.ws.column_dimensions['f'].width = 17  # 时间列
        # 所有数据居中
        max_column = self.ws.max_column
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                self.ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
        # 画折线图，定义LineChart对象
        chart = LineChart()
        # 定义x坐标的数据
        x_data = Reference(self.ws, min_col=6, min_row=2, max_col=6, max_row=max_row)
        # 定义作图数据列
        data = Reference(self.ws, min_col=7, min_row=1, max_col=7, max_row=max_row)
        # 添加数据到chart（from_rows=False说明是以列数据来画图，titles_from_data=False说明不画图例，若为True则以第一行文字写图例，这时上方的data处应min_row=1）
        chart.add_data(data, from_rows=False, titles_from_data=True)
        # 添加x坐标数据
        chart.set_categories(x_data)
        # 定义纵坐标名称
        chart.y_axis.title = '值'
        # 定义横坐标名称
        chart.x_axis.title = '时间'
        # 定义折线图标题
        chart.title = self.during_time + '-' + self.file_name
        # 把折线图画在ws表的i2处
        chart.anchor = 'i2'
        # 设置折线图大小
        chart.height = 50
        chart.width = 150
        self.ws.add_chart(chart)
        self.wb.save(f'{self.target_path}\\{self.file_name}.xlsx')

    def table_integration(self):
        # 若是第一个Excel，做一些初始化工作
        if self.first_file_key:
            # 若已存在IntegrationFile.xlsx，则覆盖
            self.integration_wb = openpyxl.Workbook()
            # 获取sheet1表
            self.integration_ws = self.integration_wb.active
        # 获取当前最大列数
        integration_max_column = self.integration_ws.max_column
        # 设置列宽

        self.integration_ws.column_dimensions[chr(integration_max_column + 1 + 65 - 1)].width = 23  # 值:联网设备列
        # 在第一行写入“时间”（需分类讨论，因为第一列有值或空，integration_max_column都是1）
        if self.integration_ws['a1'].value is None:
            self.integration_ws['a1'] = f'时间:{self.ws["a2"].value}'
            self.integration_ws.column_dimensions['a'].width = 23  # 设置列宽
        else:
            self.integration_ws.cell(row=1, column=integration_max_column + 1).value = self.ws["a2"].value
            self.integration_ws.column_dimensions[chr(integration_max_column + 1 + 65 - 1)].width = 23  # 设置列宽
        # 更新此时最大列数
        integration_max_column = self.integration_ws.max_column
        # 在第一行的最后一格追加-- 格式形如：值:霍尔2
        self.integration_ws.cell(row=1, column=integration_max_column + 1).value = self.ws["a2"].value
        self.integration_ws.column_dimensions[chr(integration_max_column + 1 + 65 - 1)].width = 23  # 设置列宽
        # 更新integration_max_column
        integration_max_column = self.integration_ws.max_column
        # 获取当前表的max_row
        integration_max_row = self.ws.max_row
        # 把时间列和值列写入integration表
        for i in range(2, integration_max_row + 1):
            # 写入时间列
            self.integration_ws.cell(row=i, column=integration_max_column - 1).value = self.ws[f'f{i}'].value
            # 写入值列
            self.integration_ws.cell(row=i, column=integration_max_column).value = self.ws[f'g{i}'].value
        # 如果此时是最后一个Excel文件，则可以开始对此表进行数据整合处理
        if self.last_file_key:
            # 时间对齐处理
            j = 2
            move_count = 0
            for i in range(2, integration_max_column, 2):
                # 以A2作为标准时间
                standard_time = self.integration_ws['a2'].value
                while True:
                    compared_time = self.integration_ws[f'{chr(65 + i)}{j}'].value
                    if standard_time != compared_time:
                        move_count += 1
                        # 清除不对齐时间、值数据
                        self.integration_ws[f'{chr(65 + i)}{j}'] = None
                        self.integration_ws[f'{chr(65 + i + 1)}{j}'] = None
                        j += 1
                    else:
                        this_column_max_row = max(
                            x.row for x in self.integration_ws[f'{chr(65 + i)}'] if x.value)
                        self.integration_ws.move_range(
                            f'{chr(65 + i)}{move_count + 2}:{chr(65 + i + 1)}{this_column_max_row}', rows=-move_count)
                        j = 2
                        move_count = 0
                        break


    def tail_data_align(self):
        self.path = input('IntegrationFile所在位置:')
        self.integration_wb = openpyxl.load_workbook(f'{self.path}\\IntegrationFile.xlsx')
        self.integration_ws = self.integration_wb.active
        self.max_column = self.integration_ws.max_column
        max_row = self.integration_ws.max_row
        # 尾部数据对齐
        for i in range(max_row, 0, -1):
            # 以最大列数为界限遍历
            for j in range(self.max_column):
                # 从最后一行看起，若有一个单元格为空，则删除此行
                if self.integration_ws[f'{chr(65 + j)}{i}'].value is None:
                    self.integration_ws.delete_rows(i)
                    break
    def make_a_difference(self):
        # 备份此时的integration_max_column
        self.tem_integration_max_column = self.integration_ws.max_column
        # 获取最大行数
        self.integration_max_row = self.integration_ws.max_row
        # 两组数据之间作差
        for i in range(2, self.tem_integration_max_column):
            for j in range(i + 1, self.tem_integration_max_column + 1):
                # 更新此时最大列数
                self.integration_max_column = self.integration_ws.max_column
                # 写入列标题：哪列减哪列的绝对值
                self.integration_ws.cell(row=1,
                                         column=self.integration_max_column + 1).value = '|' + f'{chr(65 + i - 1)}' + '-' + f'{chr(65 + j - 1)}' + '|'
                # 行遍历，获取相减结果的绝对值并写入对应单元格
                for k in range(2, self.integration_max_row + 1):
                    # 被减数
                    minuend = self.integration_ws.cell(row=k, column=i).value
                    # 减数
                    subtrahend = self.integration_ws.cell(row=k, column=j).value
                    # 若任意单元格不为空值才相减
                    if None not in (minuend, subtrahend):
                        self.integration_ws.cell(row=k, column=self.integration_max_column + 1).value = abs(
                            minuend - subtrahend)
            # 更新此时的最大列数
            self.integration_max_column += 1
            # 所有数据居中
            for i in range(1, self.integration_max_row + 1):
                for j in range(1, self.integration_max_column + 1):
                    self.integration_ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')


    def integration_mapping(self):
        # # 删除多余的时间列
        # delete_column = map(int, input('要删除的列(空格分隔):').split())
        # for i in delete_column:
        #     self.integration_ws.delete_cols(i)
        self.integration_wb.save(f'{self.path}\\IntegrationFile.xlsx')
        self.integration_wb.close()
        self.integration_max_column = self.integration_ws.max_column
        self.integration_max_row = self.integration_ws.max_row
        # 画多表整合折线图，定义LineChart对象
        chart = LineChart()
        # 定义x坐标的数据
        x_data = Reference(self.integration_ws, min_col=1, min_row=2, max_col=1, max_row=self.integration_max_row)
        # 定义作图数据列
        data = Reference(self.integration_ws, min_col=2, min_row=1, max_col=4, max_row=self.integration_max_row)
        # 添加数据到chart（from_rows=False说明是以列数据来画图，titles_from_data=False说明不画图例，若为True则以第一行文字写图例，这时上方的data处应min_row=1）
        chart.add_data(data, from_rows=False, titles_from_data=True)
        # 添加x坐标数据
        chart.set_categories(x_data)
        # 定义纵坐标名称
        chart.y_axis.title = '值'
        # 定义横坐标名称
        chart.x_axis.title = '时间'
        # 定义折线图标题
        chart.title = '数据整合'
        # 把折线图画在表的当前最大列之后处
        chart.anchor = f'{chr(self.integration_max_column + 1 + 65)}2'
        # 设置折线图大小
        chart.height = 50
        chart.width = 150
        self.integration_ws.add_chart(chart)

        # 画绝对误差折线图，定义LineChart对象
        chart = LineChart()
        x_data = Reference(self.integration_ws, min_col=1, min_row=2, max_col=1, max_row=self.integration_max_row)
        data = Reference(self.integration_ws, min_col=self.tem_integration_max_column + 1, min_row=1, max_col=self.integration_max_column, max_row=self.integration_max_row)
        chart.add_data(data, from_rows=False, titles_from_data=True)
        chart.set_categories(x_data)
        chart.y_axis.title = '差'
        chart.x_axis.title = '时间'
        chart.title = '绝对误差（取两两差的绝对值）'
        chart.anchor = f'{chr(self.integration_max_column + 1 + 65)}103'
        chart.height = 50
        chart.width = 150
        self.integration_ws.add_chart(chart)
        self.integration_wb.save(f'{self.path}\\IntegrationFile.xlsx')
        self.integration_wb.close()

    def time_disposal(self, time):
        # 时间分割
        time_lis = time.rsplit(':', 2)  # ['2023-06-05 11' , '00' , '26.000']
        time_lis_ = time_lis[0].split(' ')  # ['2023-06-05', '11']
        time_lis__ = time_lis_[0].split('-')  # ['2023', '06', '05']
        # 年月
        year_mon = '-'.join(time_lis__[:2])
        # 日
        day = time_lis__[2]
        # 时
        hour = time_lis_[1]
        # 分
        min = time_lis[1]
        # 分钟数延迟问题处理
        min = int(min) + self.time_add
        if min > 59:
            min -= 60
            hour = int(hour) + 1
            if (hour > 23):
                hour -= 24
                day = int(day) + 1
                if int(day) < 10:
                    day = '0' + str(day)
            if int(hour) < 10:
                hour = '0' + str(hour)
        if int(min) < 10:
            min = '0' + str(min)
        return year_mon + '-' + str(day) + ' ' + str(hour) + ':' + str(min)


