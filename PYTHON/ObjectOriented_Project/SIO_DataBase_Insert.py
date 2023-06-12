import openpyxl
import DataBase
import public_tools


class SIO_DataBase_Insert:
    def __init__(self):
        self.tb_staff_datas = []
        self.tb_project_admin_datas = []
        self.tb_task_datas = []
        self.project_admin_name = []
        self.task_principal_name = []
        
    def insert_project_admin_and_task(self):
        desktop_path = public_tools.My_Tool.get_desktop_path()
        title_name_lis = ['Excel文件地址:', 'Excel文件名(只支持.xlsx):', '从第几行开始:', '第几行结束:',
                          '项目名称/经理所在列:', '结果定义所在列:', '任务负责人所在列:']
        windows_name = '月任务导入数据库小工具'
        button_name = '开始导入'
        default_value_lis = [desktop_path, '1.xlsx', '4', '', 'b', 'd', 'i']
        GUI = public_tools.PySimpleGUI_Tool(title_name_lis, windows_name, button_name, default_value_lis, file_suffix=['.xlsx'], in_size=25, suffix_in_size=25)
        path, file_name, start_row, end_row, project_start_column, description_start_column, principal_start_column = GUI.gui_windows()
        wb = openpyxl.load_workbook(path + '\\' + file_name)
        ws = wb.active

        # 从Excel的标题获取表的时间作为新任务表的表名
        Excel_title = ws['a1'].value
        tb_task_title = 'tb_task_'
        for s in Excel_title:
            if s.isdigit():
                tb_task_title += s
        # 数据写入数据库的数据列表（内部包含多个元组）
        self.tb_task_datas = []
        # 项目经理列表
        self.tb_project_admin_datas = []
        # 项目经理表名
        tb_project_admin_title = 'tb_project_admin'
        # 从开始行开始遍历Excel
        for i in range(int(start_row), int(end_row) + 1):
            # 获取项目名称
            project_info = ws[f'{project_start_column}{i}'].value
            if (project_info is not None):
                if '项目经理：' in project_info:
                    # 以“：”为分隔，从右往左仅分隔一次，防止有多个“：”
                    sp = project_info.rsplit('：', 1)
                    # 有“项目经理：”字样，且被括号括住
                    if project_info[project_info.find('项目经理') - 1] in ('(', '（'):
                        project_name = sp[0][:-5]
                        self.project_admin_name = sp[1][:-1]
                    # 有“项目经理：”字样，且没有被括号括住
                    else:
                        project_name = sp[0][:-4]
                        self.project_admin_name = sp[1]
                # 没有“项目经理：”字样，只写了个名字，但名字被括号括住的情况
                elif '（' in project_info:
                    sp = project_info.rsplit('（', 1)
                    project_name = sp[0]
                    self.project_admin_name = sp[1][:-1]
                else:
                    project_name = project_info
                    self.project_admin_name = None
                # 去除项目名字内的换行符
                project_name_ = ''
                for s in project_name:
                    if s not in ('\n', '\xa0'):
                        project_name_ += s
                # 去除项目名字最后的句号
                if (project_name_[-1] == '。'):
                    project_name_ = project_name_[:-1]
                # 去除项目名字后面可能多余的空格
                project_name_ = project_name_.strip()
            # 获取结果定义
            task_description = ws[f'{description_start_column}{i}'].value
            # 若该格结果定义非空，去除结果定义中可能的换行符和前后的空格
            if task_description is not None:
                task_description = ''.join(task_description.split('\n')).strip()
                # 去除结果定义最后可能存在的句号
                if task_description[-1] == '。':
                    task_description = task_description[:-1]
            # 获取该结果定义的负责人
            self.task_principal_name = ws[f'{principal_start_column}{i}'].value
            # 把负责人名字唯一追加到员工表
            if self.task_principal_name not in self.tb_staff_datas and self.task_principal_name is not None:
                self.tb_staff_datas.append(self.task_principal_name)
            # 把tb_task数据追加到任务列表中
            self.tb_task_datas.append((project_name_, self.project_admin_name, task_description, self.task_principal_name))
            if self.project_admin_name is not None:
                # 把项目经理名字唯一追加到项目经理列表中
                if self.project_admin_name not in self.tb_project_admin_datas:
                    self.tb_project_admin_datas.append(self.project_admin_name)
                # 把项目经理名字唯一追加到员工表
                if self.project_admin_name not in self.tb_staff_datas:
                    self.tb_staff_datas.append(self.project_admin_name)
        # 关闭Excel文件
        wb.close()

        # 创建数据库类对象
        db = DataBase.DataBase(host='localhost')
        '''
            插入任务表数据
        '''
        # 若存在当月的任务计划表，则先删除（预防重新导入的可能）
        db.drop_table_if_exists(tb_task_title)
        # 若不存在当月的任务计划表，则创建
        db.create_task_table_if_not_exists(tb_task_title)
        # 遍历self.tb_task_datas，获取每行元祖数据并插入
        for task_data in self.tb_task_datas:
            db.insert(tb_task_title, task_data, data_num=4)
        '''
            插入当月任务表的项目经理（若已存在，则不插入）
        '''
        # 遍历self.tb_project_admin_datas，获取每个项目经理字符串插入
        for project_admin_data in self.tb_project_admin_datas:
            if db.select_tb_project_admin(project_admin_data) == 1:
                db.insert(tb_project_admin_title, project_admin_data, data_num=1)
        # 提交事务
        db.commit_and_close()

    '''
        从Excel中引入员工姓名到tb_staff，若员工已存在则不添加
    '''
    def insert_staff(self):
        # 从上方的月任务表引入员工信息
        if self.tb_staff_datas is not []:
            db = DataBase.DataBase(host='localhost')
            db.create_staff_if_not_exists()
            for staff_data in self.tb_staff_datas:
                if db.select_tb_staff(staff_data) == 1:
                    db.insert_staff(staff_data)
            db.commit_and_close()
        # 从另一个Excel引入员工信息
        else:
            pass
if __name__ == '__main__':
    ins = SIO_DataBase_Insert()
    # 从Excel月任务表中不重复地导入项目经理构成项目经理表。且把相应数据创建月任务表且导入，表根据Excel标题动态命名为例如：tb_task_2305
    ins.insert_project_admin_and_task()
    # 从Excel月任务表中导入员工（不管是负责人还是项目经理，只要是员工姓名且未被导入过就导入，不会重复导入）
    ins.insert_staff()